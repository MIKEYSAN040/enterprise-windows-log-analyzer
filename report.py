"""
report.py
----------
Turns the list of finding dicts produced by the detectors into a formatted
Excel workbook - one sheet per detection category, plus a summary sheet.
Same OpenPyXL pattern used in the Cloud Access Privilege Auditor project.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SEVERITY_COLORS = {
    "High": PatternFill(start_color="F8696B", end_color="F8696B", fill_type="solid"),
    "Medium": PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid"),
    "Low": PatternFill(start_color="63BE7B", end_color="63BE7B", fill_type="solid"),
}


def _write_sheet(ws, findings):
    if not findings:
        ws.append(["No findings in this category"])
        return

    # Union of all keys across findings becomes the column set, MITRE flattened out
    columns = []
    for f in findings:
        for k in f.keys():
            if k == "mitre":
                for mk in ("technique_id", "technique_name", "tactic"):
                    col = f"mitre_{mk}"
                    if col not in columns:
                        columns.append(col)
            elif k not in columns:
                columns.append(k)

    ws.append(columns)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for f in findings:
        row = []
        for col in columns:
            if col.startswith("mitre_"):
                mitre_key = col.replace("mitre_", "")
                value = f.get("mitre", {}).get(mitre_key, "")
            else:
                value = f.get(col, "")
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            row.append(str(value) if value is not None else "")
        ws.append(row)

        if "severity" in columns:
            sev_col = columns.index("severity") + 1
            sev_val = row[columns.index("severity")]
            fill = SEVERITY_COLORS.get(sev_val)
            if fill:
                ws.cell(row=ws.max_row, column=sev_col).fill = fill

    for i, col in enumerate(columns, start=1):
        max_len = max([len(col)] + [len(str(r[i - 1])) for r in ws.iter_rows(min_row=2, values_only=True)])
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)


def generate_report(results: dict, output_path: str):
    """
    results: dict like {"brute_force": [...], "privilege_escalation": [...], ...}
    """
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append(["Detection Category", "Findings Count", "MITRE Technique"])
    for cell in summary_ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for category, findings in results.items():
        technique = findings[0]["mitre"]["technique_id"] if findings else "-"
        summary_ws.append([category.replace("_", " ").title(), len(findings), technique])

    for category, findings in results.items():
        ws = wb.create_sheet(title=category[:31])
        _write_sheet(ws, findings)

    wb.save(output_path)
    return output_path
